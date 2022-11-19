from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from django.shortcuts import render
from django.shortcuts import redirect
from django.views.generic import ListView, CreateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from dashboard.forms import job_industry_form, total_records_form, job_profile_form
from django.http import JsonResponse
from dashboard.models import Job_industries, job_profiles, Posts
from dashboard.forms import job_industry_form, total_records_form, blog_form
from django.http import JsonResponse
from dashboard.models import Job_industries
from django.contrib.auth.models import User
import openpyxl


def login_success(request):
    """
    Redirects users based on whether they are in the admins group
    """
    id = request.user.id
    print("id", id)
    print(request.user.username)
    return redirect("home_page")


class dashboard_home(LoginRequiredMixin, CreateView):
    """
    Summary line.
    This class Show the admin dashboard

    """

    model = Job_industries
    template_name = "public/record.html"
    form_class = total_records_form

    def get_form(self, form_class=total_records_form):
        form = super(dashboard_home, self).get_form(form_class)
        form.fields["total_industry"].initial = Job_industries.objects.filter(
            status=1
        ).count()
        form.fields["total_blogs"].initial = Posts.objects.filter(status=1).count()
        form.fields["total_jobs"].initial = job_profiles.objects.filter(
            status=1
        ).count()
        form.fields["total_user"].initial = User.objects.filter(is_active=1).count()
        return form


@login_required
def create_job_industries(request):
    form = job_industry_form()
    available_industries = Job_industries.objects.filter(status=1).values_list(
        "name", "id"
    )
    data = []
    for i in range(len(available_industries)):
        data.append(
            [
                i,
                str(available_industries[i][0]),
                '<a class="btn btn-info btn-sm" values=('
                + str(available_industries[i])
                + ') onclick="edit('
                + str(available_industries[i][1])
                + ')">'
                + "Edit"
                + "</a>",
                '<a class="btn btn-info btn-sm"  onclick="delete_industry('
                + str(available_industries[i][1])
                + ')">'
                + "Delete"
                + "</a>",
            ]
        )
    return render(
        request, "public/new.html", {"form": form, "available_industries": data}
    )


@login_required
def create_blog_post(request):
    form = blog_form()
    available_industries = Posts.objects.filter(status=1).values_list(
        "id", "title", "author", "slug"
    )
    data = []
    for i in range(len(available_industries)):
        data.append(
            [
                i,
                str(available_industries[i][1]),
                str(available_industries[i][2]),
                str(available_industries[i][3]),
                '<a class="btn btn-info btn-sm" values=('
                + str(available_industries[i][0])
                + ') onclick="edit('
                + str(available_industries[i][0])
                + ')">'
                + "Edit"
                + "</a>",
                '<a class="btn btn-info btn-sm"  onclick="delete_blog('
                + str(available_industries[i][0])
                + ')">'
                + "Delete"
                + "</a>",
            ]
        )
    return render(
        request, "public/blog.html", {"form": form, "available_blog": data}
    )


@login_required
def add_industry(request):
    industry = request.GET.get("industry_name")
    data = {"added": False}
    print(industry, request.user.id)
    if not Job_industries.objects.filter(name=industry).exists():
        print(Job_industries.objects.filter(name=industry).exists())
        add_industry_obj = Job_industries(
            name=industry,
            status=1,
            updated_by_user_id=int(request.user.id),
            created_on=datetime.now(),
            updated_on=datetime.now(),
        )
        add_industry_obj.save()
        data["added"] = True
    return JsonResponse(data)


@login_required
def edit_job_industries(request):
    form = job_industry_form()
    return render(request, "public/update_industry.html", {"form": form})


@login_required
def edit_job_profile(request):
    form = job_profile_form()
    return render(request, "public/update_job_profile.html", {"form": form})


@login_required
def update_job_industries(request):
    id = request.GET.get("id")
    data = {"added": False}
    find_data = Job_industries.objects.filter(id=id).exists()
    find_data1 = Job_industries.objects.filter(name=request.GET.get("name")).exists()
    if find_data1:
        data["added"] = False
        return JsonResponse(data)
    if find_data:
        find_data = Job_industries.objects.filter(id=id)
        add_industry_objet = find_data.update(
            name=request.GET.get("name"), updated_on=datetime.now()
        )
        data["added"] = True
    return JsonResponse(data)


@login_required
def update_job_profile(request):
    id = request.GET.get("id")
    industry = request.GET.get("companey_name")
    job_Title = request.GET.get("job_Title")
    JobType = request.GET.get("job_type")
    job_Category = request.GET.get("job_Category")
    job_skill = request.GET.get("job_skill")
    gender = request.GET.get("gender")
    job_url = request.GET.get("job_url")
    mobile_no = request.GET.get("mobile_no")
    email = request.GET.get("email")
    salary = request.GET.get("salary")
    country = request.GET.get("country")
    state = request.GET.get("state")
    city = request.GET.get("city")
    career_level = request.GET.get("career_level")
    degree_level = request.GET.get("degree_level")
    job_experience = request.GET.get("job_experience")
    description = request.GET.get("description")
    remote_work = request.GET.get("remote_work")
    job_exp_date = str(date.today() + relativedelta(months=+3))
    data = {"added": False}
    find_data = job_profiles.objects.filter(id=id)
    print(job_url)
    if find_data.exists():
        add_industry_objet = find_data.update(
            job_type=JobType,
            email=email,
            mobile_no=mobile_no,
            job_url=job_url,
            companey_name=industry,
            status=1,
            job_Title=job_Title,
            job_Category=job_Category,
            job_skill=job_skill,
            gender=gender,
            job_exp_date=job_exp_date,
            salary_package=salary,
            country=country,
            state=state,
            city=city,
            career_level=career_level,
            degree_level=degree_level,
            job_experience=job_experience,
            description=description,
            remote_work=remote_work,
            updated_by_user_id=request.user.id,
            created_on=datetime.now(),
            updated_on=datetime.now(),
        )
        data["added"] = True
    return JsonResponse(data)


@login_required
def delete_industry(request):
    id = request.POST.get("id")
    data = {"is_taken": False}
    find_data = Job_industries.objects.filter(id=id)
    if find_data.exists():
        add_industry_objet = find_data.update(status=0, updated_on=datetime.now())
        data["is_taken"] = True
    return JsonResponse(data)


@login_required
def delete_blog_post(request):
    id = request.POST.get("id")
    data = {"is_taken": False}
    find_data = Posts.objects.filter(id=id)
    if find_data.exists():
        add_industry_objet = find_data.update(status=0, updated_on=datetime.now())
        data["is_taken"] = True
    return JsonResponse(data)


@login_required
def delete_job_profile(request):
    id = request.POST.get("id")
    data = {"is_taken": False}
    find_data = job_profiles.objects.filter(id=id)
    if find_data.exists():
        add_industry_objet = find_data.update(status=0, updated_on=datetime.now())
        data["is_taken"] = True
    return JsonResponse(data)


@login_required
def create_job_profile(request):
    form = job_profile_form()
    data = []
    available_industries = job_profiles.objects.filter(status=1).values_list(
        "id",
        "companey_name",
        "job_Title",
        "job_Category",
        "job_exp_date",
        "updated_by_user_id",
    )
    for i in range(len(available_industries)):
        data.append(
            [
                i,
                str(available_industries[i][1]),
                str(available_industries[i][2]),
                str(available_industries[i][3]),
                str(available_industries[i][4]),
                str(available_industries[i][5]),
                '<a class="btn btn-info btn-sm" values=('
                + str(available_industries[i])
                + ') onclick="edit('
                + str(available_industries[i][0])
                + ')">'
                + "Edit"
                + "</a>",
                '<a class="btn btn-info btn-sm"  onclick="delete_job_profile('
                + str(available_industries[i][0])
                + ')">'
                + "Delete"
                + "</a>",
            ]
        )
    return render(
        request, "public/job_profile.html", {"form": form, "available_industries": data}
    )


@login_required
def get_job_name(request):
    id = request.GET.get("id")
    get_data = Job_industries.objects.filter(id=id).first()
    data = {"added": get_data.name}
    return JsonResponse(data)


@login_required
def get_job_profile(request):
    id = request.GET.get("id")
    get_data = job_profiles.objects.filter(id=id).first()
    print(
        get_data.companey_name,
        "ddddddddddddddddddddddddddddddddddddddddddddddddddddddd",
    )
    data = {
        "companey_name": get_data.companey_name,
        "job_Title": get_data.job_Title,
        "job_type": get_data.job_type,
        "job_Category": get_data.job_Category,
        "job_skill": get_data.job_skill,
        "gender": get_data.gender,
        "salary_package": get_data.salary_package,
        "country": get_data.country,
        "state": get_data.state,
        "city": get_data.city,
        "career_level": get_data.career_level,
        "degree_level": get_data.degree_level,
        "job_experience": get_data.job_experience,
        "description": get_data.description,
        "remote_work": get_data.remote_work,
        "job_url": get_data.job_url,
        "email": get_data.email,
        "mobile_no": get_data.mobile_no,
    }
    return JsonResponse(data)


@login_required
def add_job_profile(request):
    industry = request.GET.get("companey_name")
    job_Title = request.GET.get("job_Title")
    JobType = request.GET.get("job_type")
    job_Category = request.GET.get("job_Category")
    job_skill = request.GET.get("job_skill")
    gender = request.GET.get("gender")
    job_url = request.GET.get("job_url")
    mobile_no = request.GET.get("mobile_no")
    email = request.GET.get("email")
    salary = request.GET.get("salary")
    country = request.GET.get("country")
    state = request.GET.get("state")
    city = request.GET.get("city")
    career_level = request.GET.get("career_level")
    degree_level = request.GET.get("degree_level")
    job_experience = request.GET.get("job_experience")
    description = request.GET.get("description")
    remote_work = request.GET.get("remote_work")
    data = {"added": False}
    job_exp_date = str(date.today() + relativedelta(months=+3))
    add_industry_obj = job_profiles(
        job_type=JobType,
        email=email,
        mobile_no=mobile_no,
        job_url=job_url,
        companey_name=industry,
        status=1,
        job_Title=job_Title,
        job_Category=job_Category,
        job_skill=job_skill,
        gender=gender,
        job_exp_date=job_exp_date,
        salary_package=salary,
        country=country,
        state=state,
        city=city,
        career_level=career_level,
        degree_level=degree_level,
        job_experience=job_experience,
        description=description,
        remote_work=remote_work,
        updated_by_user_id=request.user.id,
        created_on=datetime.now(),
        updated_on=datetime.now(),
    )
    add_industry_obj.save()
    data["added"] = True
    return JsonResponse(data)


@login_required
def add_blog_post(request):
    title = request.GET.get("title")
    content = request.GET.get("content")
    image = request.GET.get("image")
    slug = request.GET.get("slug")
    data = {"added": False}
    add_industry_obj = Posts(
        title=title,
        blog_image=image,
        content=content,
        status=1,
        slug=slug,
        author_id=request.user.id,
        created_on=datetime.now(),
        updated_on=datetime.now(),
    )
    add_industry_obj.save()
    data["added"] = True
    return JsonResponse(data)


@login_required
def show_user(request):
    available_users = User.objects.filter(is_active=1).values_list(
        "id", "username", "first_name", "last_name", "email", "date_joined"
    )
    data = []
    for i in range(len(available_users)):
        data.append(
            [
                available_users[i][0],
                str(available_users[i][1]),
                str(available_users[i][2]),
                str(available_users[i][3]),
                str(available_users[i][4]),
                str(available_users[i][5]),
                '<a class="btn btn-info btn-sm" onclick="edit('
                + str(available_users[i][0])
                + ')">'
                + "Edit"
                + "</a>",
                '<a class="btn btn-info btn-sm"  onclick="delete_user('
                + str(available_users[i][0])
                + ')">'
                + "Delete"
                + "</a>",
            ]
        )
    return render(request, "public/user.html", {"available_users": data})


@login_required
def delete_user(request):
    id = request.GET.get("id")
    data = {"is_taken": False}
    find_user = User.objects.filter(id=id)
    if find_user.exists():
        delete_user = find_user.update(is_active=0)
        data["is_taken"] = True
    return JsonResponse(data)


# Pending Not Perfect working
@login_required
def index(request):
    form = blog_form()
    excel_file = request.FILES["excel_file"]

    # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file)

    # getting all sheets
    sheets = wb.sheetnames
    excel_data = list()

    for sheet in sheets:

        # getting a particular sheet
        worksheet = wb[sheet]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)
        row_data = ['Title' ,'Slug' ,'Content'  ,'Blog Image']
        excel_row_data =[]
        for row in list(worksheet[1]):
                excel_row_data.append(str(row.value))
        print(excel_row_data)
        if excel_row_data!=row_data:
            return render(request, 'public/blog.html', {"errors": 'Upload Right Excel Format',"form": form,"available_blog":[]})
 
        # iterating over the rows and
        # getting value from each cell in row
        co=0
        for row in worksheet.iter_rows():
            row_data = list()
            if co!=0:
                for cell in row:
                    row_data.append(str(cell.value))
                    # print(cell.value)
                excel_data.append(row_data)
            co+=1
    return render(request, "public/blog.html", {"available_blog": excel_data,"form": form})
